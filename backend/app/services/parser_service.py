from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from app.core.enums import AmountMode, TransactionDirection
from app.schemas.standard import StandardBankTransaction

HEADER_KEYWORDS = {
    "交易日期",
    "入账日期",
    "记账日期",
    "收入",
    "支出",
    "借方发生额",
    "贷方发生额",
    "金额",
    "余额",
    "对方户名",
    "对方账号",
    "对方银行",
    "摘要",
    "附言",
    "用途",
    "流水号",
    "交易流水号",
}


@dataclass(slots=True)
class BankTemplateParseConfig:
    bank_account_id: str
    source_file_id: str
    file_type: str
    sheet_name: str
    header_row_index: int
    data_start_row_index: int
    field_aliases: dict[str, str]
    amount_mode: AmountMode
    amount_config: dict[str, str]
    date_formats: list[str]
    currency: str = "CNY"


CellValue = str | date | datetime | int | float | Decimal | None


def detect_header_row(rows: list[list[CellValue]], scan_limit: int = 30) -> int:
    best_index = 0
    best_score = -1

    for index, row in enumerate(rows[:scan_limit]):
        score = 0
        for cell in row:
            value = _clean_cell(cell)
            if not value:
                continue
            if value in HEADER_KEYWORDS:
                score += 2
            if any(keyword in value for keyword in HEADER_KEYWORDS):
                score += 1

        if score > best_score:
            best_index = index
            best_score = score

    return best_index


def parse_bank_statement(
    path: str | Path,
    config: BankTemplateParseConfig,
) -> list[StandardBankTransaction]:
    rows = _read_rows(path, config.file_type, config.sheet_name)
    if config.header_row_index >= len(rows):
        raise ValueError("Header row index is out of range")

    header_row = rows[config.header_row_index]
    transactions: list[StandardBankTransaction] = []

    for row_index in range(config.data_start_row_index, len(rows)):
        row = rows[row_index]
        if _row_is_empty(row):
            continue

        raw_row = _build_raw_row(header_row, row)
        normalized_row = _normalize_row(raw_row, config.field_aliases)
        if not normalized_row:
            continue

        transaction_date = _require_value(normalized_row, "transaction_date")
        posting_date = normalized_row.get("posting_date")
        direction, debit_amount, credit_amount, net_amount = _parse_amounts(
            normalized_row,
            config.amount_mode,
            config.amount_config,
        )

        balance = _decimal_or_none(normalized_row.get("balance"))
        transactions.append(
            StandardBankTransaction(
                transaction_date=_parse_date(transaction_date, config.date_formats),
                posting_date=_parse_optional_date(posting_date, config.date_formats),
                bank_account_id=config.bank_account_id,
                currency=config.currency,
                direction=direction,
                debit_amount=debit_amount,
                credit_amount=credit_amount,
                net_amount=net_amount,
                balance=balance,
                counterparty_name=_none_if_blank(normalized_row.get("counterparty_name")),
                counterparty_account_no=_none_if_blank(normalized_row.get("counterparty_account_no")),
                counterparty_bank_name=_none_if_blank(normalized_row.get("counterparty_bank_name")),
                summary=_none_if_blank(normalized_row.get("summary")),
                purpose=_none_if_blank(normalized_row.get("purpose")),
                transaction_type=_none_if_blank(normalized_row.get("transaction_type")),
                bank_transaction_id=_none_if_blank(normalized_row.get("bank_transaction_id")),
                receipt_no=_none_if_blank(normalized_row.get("receipt_no")),
                source_file_id=config.source_file_id,
                source_sheet_name=config.sheet_name,
                source_row_index=row_index + 1,
                raw_row=raw_row,
            )
        )

    return transactions


def _read_rows(path: str | Path, file_type: str, sheet_name: str) -> list[list[CellValue]]:
    file_path = Path(path)
    normalized_type = file_type.lower()

    if normalized_type == "csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [[_clean_cell(cell) for cell in row] for row in csv.reader(handle)]

    if normalized_type == "xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: {sheet_name}")
            sheet = workbook[sheet_name]
            return [
                list(row)
                for row in sheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()

    if normalized_type == "xls":
        raise ValueError("Unsupported file type: xls")

    raise ValueError(f"Unsupported file type: {file_type}")


def _build_raw_row(header_row: list[CellValue], row: list[CellValue]) -> dict[str, CellValue]:
    raw_row: dict[str, CellValue] = {}

    for index, header in enumerate(header_row):
        key = _clean_cell(header)
        if not key:
            continue
        raw_row[key] = row[index] if index < len(row) else None

    return raw_row


def _normalize_row(
    raw_row: dict[str, CellValue],
    field_aliases: dict[str, str],
) -> dict[str, CellValue]:
    normalized: dict[str, CellValue] = {}

    for header, value in raw_row.items():
        canonical_field = field_aliases.get(header)
        if canonical_field is None:
            continue
        normalized[canonical_field] = value

    return normalized


def _parse_amounts(
    normalized_row: dict[str, CellValue],
    amount_mode: AmountMode,
    amount_config: dict[str, str],
) -> tuple[TransactionDirection, Decimal | None, Decimal | None, Decimal]:
    if amount_mode != AmountMode.INCOME_EXPENSE_COLUMNS:
        raise ValueError(f"Unsupported amount mode: {amount_mode}")

    income = _decimal_or_none(normalized_row.get(amount_config["income"]))
    expense = _decimal_or_none(normalized_row.get(amount_config["expense"]))

    if (
        income is not None
        and expense is not None
        and income != Decimal("0")
        and expense != Decimal("0")
    ):
        raise ValueError("Both income and expense amounts are populated")

    if income is not None and income != Decimal("0"):
        return (
            TransactionDirection.CREDIT,
            None,
            income,
            income,
        )

    if expense is not None and expense != Decimal("0"):
        return (
            TransactionDirection.DEBIT,
            expense,
            None,
            -expense,
        )

    raise ValueError("Unable to determine transaction amount")


def _parse_date(value: CellValue, date_formats: list[str]) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    candidate = _clean_cell(value)
    for date_format in date_formats:
        try:
            return datetime.strptime(candidate, date_format).date().isoformat()
        except ValueError:
            continue

    raise ValueError(f"Invalid date: {value}")


def _parse_optional_date(value: CellValue, date_formats: list[str]) -> str | None:
    if value is None or not _clean_cell(value):
        return None
    return _parse_date(value, date_formats)


def _decimal_or_none(value: CellValue) -> Decimal | None:
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    candidate = _clean_cell(value).replace(",", "")
    if not candidate:
        return None

    try:
        return Decimal(candidate)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid amount: {value}") from exc


def _require_value(row: dict[str, CellValue], key: str) -> CellValue:
    value = row.get(key)
    if value is None or not _clean_cell(value):
        raise ValueError(f"Missing required field: {key}")
    return value


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _none_if_blank(value: CellValue) -> str | None:
    if value is None:
        return None
    cleaned = _clean_cell(value)
    return cleaned or None


def _row_is_empty(row: list[CellValue]) -> bool:
    return all(not _clean_cell(cell) for cell in row)
