from __future__ import annotations

import csv
import io
import itertools
from collections.abc import Iterator
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.tools.bank_journal.domain.amounts import AmountError, SignedAmount
from app.tools.bank_journal.enums import AmountMode, ExceptionCode, TransactionDirection
from app.tools.bank_journal.schemas.standard import StandardBankTransaction

HEADER_KEYWORDS = {
    "交易日期",
    "入账日期",
    "记账日期",
    "收入",
    "支出",
    "借方发生额",
    "贷方发生额",
    "金额",
    "余额",
    "对方户名",
    "对方账号",
    "对方银行",
    "摘要",
    "附言",
    "用途",
    "流水号",
    "交易流水号",
}

# 日记账表头关键词（用于日记账模板 detect：识别哪一行是表头）。
# 与银行流水关键词不重叠——日记账样本的表头通常是"凭证号/日期/科目/借方/贷方"等。
JOURNAL_HEADER_KEYWORDS = {
    "凭证号",
    "凭证字号",
    "日期",
    "摘要",
    "科目",
    "科目代码",
    "科目名称",
    "借方",
    "贷方",
    "金额",
    "余额",
    "对方",
    "备注",
    "制单",
    "附件",
}

# 日记账惯例必填列（detect 推断默认必填，仅保留与识别列名的交集）。
JOURNAL_DEFAULT_REQUIRED = {"日期", "科目", "金额"}

# 表头中文关键词 → 标准字段名（用于自动识别字段别名）。
HEADER_FIELD_MAP: dict[tuple[str, ...], str] = {
    ("交易日期", "记账日期", "日期"): "transaction_date",
    ("入账日期",): "posting_date",
    ("收入",): "income_amount",
    ("支出",): "expense_amount",
    ("借方发生额", "借方金额", "借方"): "debit_amount",
    ("贷方发生额", "贷方金额", "贷方"): "credit_amount",
    ("金额", "发生额", "交易金额"): "amount",
    ("方向", "收支方向"): "direction",
    ("余额", "账户余额"): "balance",
    ("对方户名", "对方名称", "对方账户名称"): "counterparty_name",
    ("对方账号", "对方账户", "对方账户账号"): "counterparty_account_no",
    ("对方开户行", "对方银行"): "counterparty_bank_name",
    ("摘要", "摘要信息"): "summary",
    ("用途", "附言"): "purpose",
    ("交易类型", "业务类型"): "transaction_type",
    ("流水号", "交易流水号", "交易号"): "bank_transaction_id",
    ("回单号",): "receipt_no",
}


@dataclass(slots=True)
class BankTemplateParseConfig:
    bank_account_id: str
    source_file_id: str
    file_type: str
    sheet_name: str
    header_row_index: int
    data_start_row_index: int
    field_aliases: dict[str, str]
    amount_mode: AmountMode
    amount_config: dict[str, str]
    date_formats: list[str]
    currency: str = "CNY"
    # 公司级自定义扩展字段定义。解析时据此填充 extra_fields。
    # service 层不依赖 ORM 模型，用轻量 dataclass 承载。
    custom_fields: list[CustomFieldDef] = field(default_factory=list)


@dataclass(slots=True)
class CustomFieldDef:
    """公司级自定义扩展字段的解析期定义（与 ORM CustomField 解耦的轻量视图）。"""

    field_key: str
    slot_key: str
    data_type: str  # text / amount / date
    header_keywords: list[str]


@dataclass(slots=True)
class ParsedBankRow:
    """单行解析结果。成功时 transaction 有值；失败时 parse_errors 非空。

    两者互斥：解析失败不产生 transaction，但保留 source_row_index / raw_row
    用于追溯，并标记异常码进入待人工确认（PRD §6.8 逐行识别异常）。
    """

    source_file_id: str
    source_sheet_name: str
    source_row_index: int
    raw_row: dict[str, CellValue]
    transaction: StandardBankTransaction | None = None
    parse_errors: list[ExceptionCode] = field(default_factory=list)
    error_message: str | None = None
    warnings: list[ExceptionCode] = field(default_factory=list)


CellValue = str | date | datetime | int | float | Decimal | None


def detect_header_row(
    rows: list[list[CellValue]],
    scan_limit: int = 30,
    keywords: set[str] | None = None,
) -> int:
    """扫描前 scan_limit 行，按关键词命中数打分，返回最可能是表头行的下标。

    keywords 缺省时用银行流水关键词 HEADER_KEYWORDS（保持向后兼容）；
    日记账模板 detect 传入 JOURNAL_HEADER_KEYWORDS。
    """
    effective_keywords = keywords if keywords is not None else HEADER_KEYWORDS
    best_index = 0
    best_score = -1

    for index, row in enumerate(rows[:scan_limit]):
        score = 0
        for cell in row:
            value = _clean_cell(cell)
            if not value:
                continue
            if value in effective_keywords:
                score += 2
            if any(keyword in value for keyword in effective_keywords):
                score += 1

        if score > best_score:
            best_index = index
            best_score = score

    return best_index


def parse_bank_rows(
    path: str | Path,
    config: BankTemplateParseConfig,
) -> list[ParsedBankRow]:
    """解析银行流水，逐行返回结果。流式处理，不全量物化源行列表。

    与 `parse_bank_statement` 不同，本函数**不会因单行解析错误中断**：
    日期/金额/方向无法解析的行会产出带异常码的 `ParsedBankRow`（transaction
    为 None），其余行继续解析。这符合 PRD §6.8「逐行识别异常」的设计。

    仅以下结构性错误会抛出（与单行数据无关）：
    - 表头行越界、Sheet 不存在、不支持的文件类型。
    """
    header_row: list[CellValue] | None = None
    results: list[ParsedBankRow] = []

    for row_index, row in enumerate(_read_rows(path, config.file_type, config.sheet_name)):
        if row_index == config.header_row_index:
            header_row = row
            continue
        if row_index < config.data_start_row_index:
            continue
        # 若已越过 data_start 但尚未见到 header，配置越界
        if header_row is None:
            raise ValueError("Header row index is out of range")
        if _row_is_empty(row):
            continue

        raw_row = _build_raw_row(header_row, row)
        normalized_row = _normalize_row(raw_row, config.field_aliases)
        if not normalized_row:
            continue

        parsed = _try_parse_row(normalized_row, raw_row, row_index + 1, config)
        results.append(parsed)

    # 文件行数不足以到达 header_row_index
    if header_row is None:
        raise ValueError("Header row index is out of range")

    return results


def _try_parse_row(
    normalized_row: dict[str, CellValue],
    raw_row: dict[str, CellValue],
    source_row_index: int,
    config: BankTemplateParseConfig,
) -> ParsedBankRow:
    """尝试解析单行。失败时收集异常码但不抛出。"""
    base = ParsedBankRow(
        source_file_id=config.source_file_id,
        source_sheet_name=config.sheet_name,
        source_row_index=source_row_index,
        raw_row=raw_row,
    )

    # transaction_date 缺失 → MISSING_REQUIRED_FIELD
    transaction_date_raw = normalized_row.get("transaction_date")
    if transaction_date_raw is None or not _clean_cell(transaction_date_raw):
        base.parse_errors.append(ExceptionCode.MISSING_REQUIRED_FIELD)
        base.error_message = "Missing required field: transaction_date"
        return base

    # 日期解析
    try:
        transaction_date = _parse_date(transaction_date_raw, config.date_formats)
    except ValueError:
        base.parse_errors.append(ExceptionCode.INVALID_DATE)
        base.error_message = f"Invalid date: {transaction_date_raw}"
        return base

    posting_date = normalized_row.get("posting_date")
    try:
        posting_date = _parse_optional_date(posting_date, config.date_formats)
    except ValueError:
        base.parse_errors.append(ExceptionCode.INVALID_DATE)
        base.error_message = f"Invalid posting date: {posting_date}"
        return base

    # 金额/方向解析
    try:
        amount = _parse_amounts(normalized_row, config.amount_mode, config.amount_config)
    except ValueError as exc:
        message = str(exc)
        base.parse_errors.append(ExceptionCode.INVALID_AMOUNT)
        if "direction" in message.lower():
            base.parse_errors[-1] = ExceptionCode.UNKNOWN_DIRECTION
        base.error_message = message
        return base

    if amount.sign_anomaly:
        base.warnings.append(ExceptionCode.AMOUNT_DIRECTION_MISMATCH)

    balance = _decimal_or_none(normalized_row.get("balance"))

    # 公司级自定义扩展字段：从 normalized_row 取值，按 data_type 解析后填入 extra_fields。
    # _detect_field_aliases 已把扩展字段表头映射为 field_key，故这里用 field_key 取值。
    extra_fields: dict[str, Any] = {}
    for cf in config.custom_fields:
        raw = normalized_row.get(cf.field_key)
        cleaned = _clean_cell(raw) if raw is not None else None
        if not cleaned:
            continue
        try:
            if cf.data_type == "amount":
                extra_fields[cf.field_key] = _decimal_or_none(cleaned)
            elif cf.data_type == "date":
                extra_fields[cf.field_key] = _parse_date(cleaned, config.date_formats)
            else:  # text
                extra_fields[cf.field_key] = str(cleaned)
        except ValueError:
            # 扩展字段解析失败不阻断整行（保守：核心字段已成功），仅跳过该字段
            continue

    base.transaction = StandardBankTransaction(
        transaction_date=transaction_date,
        posting_date=posting_date,
        bank_account_id=config.bank_account_id,
        currency=config.currency,
        direction=amount.direction,
        debit_amount=amount.debit_amount,
        credit_amount=amount.credit_amount,
        net_amount=amount.net_amount,
        balance=balance,
        counterparty_name=_none_if_blank(normalized_row.get("counterparty_name")),
        counterparty_account_no=_none_if_blank(normalized_row.get("counterparty_account_no")),
        counterparty_bank_name=_none_if_blank(normalized_row.get("counterparty_bank_name")),
        summary=_none_if_blank(normalized_row.get("summary")),
        purpose=_none_if_blank(normalized_row.get("purpose")),
        transaction_type=_none_if_blank(normalized_row.get("transaction_type")),
        bank_transaction_id=_none_if_blank(normalized_row.get("bank_transaction_id")),
        receipt_no=_none_if_blank(normalized_row.get("receipt_no")),
        extra_fields=extra_fields,
        source_file_id=config.source_file_id,
        source_sheet_name=config.sheet_name,
        source_row_index=source_row_index,
        raw_row=raw_row,
    )
    return base


def parse_bank_statement(
    path: str | Path,
    config: BankTemplateParseConfig,
) -> list[StandardBankTransaction]:
    """向后兼容封装：仅返回解析成功的交易（跳过失败行）。

    新代码应直接使用 `parse_bank_rows` 以获得逐行异常码。
    """
    return [row.transaction for row in parse_bank_rows(path, config) if row.transaction]


def detect_bank_template_config(
    path: str | Path,
    file_type: str,
    sheet_name: str = "",
    custom_field_defs: list[CustomFieldDef] | None = None,
    builtin_keyword_overrides: dict[str, list[str]] | None = None,
) -> dict[str, object]:
    """从样本文件自动识别银行模板配置（PRD §5.1.3 / §9.1）。

    识别：表头行位置、数据起始行、字段别名、金额模式、日期格式候选。
    ``custom_field_defs`` 传入公司级扩展字段后，表头识别会包含扩展字段。
    ``builtin_keyword_overrides`` 传入公司级内置关键词覆盖后，内置字段按覆盖集识别。
    返回可直接填入 `BankTemplateVersionCreate` 的 dict。
    """
    resolved_sheet = sheet_name or _first_sheet_name(path, file_type)
    it = _read_rows(path, file_type, resolved_sheet)
    # 表头检测只需前 scan_limit 行；chain 回链剩余，无需重开文件
    scan_limit = 30
    window = list(itertools.islice(it, scan_limit))
    header_row_index = detect_header_row(window)
    # 模板检测需随机访问完整行列表（样本文件，物化可接受）
    rows = list(itertools.chain(window, it))
    header_row = rows[header_row_index] if header_row_index < len(rows) else []
    data_start_row_index = _detect_data_start_row(rows, header_row_index)

    field_aliases = _detect_field_aliases(
        header_row, custom_field_defs, builtin_keyword_overrides
    )
    amount_mode, amount_config = _detect_amount_mode(field_aliases)
    date_formats = _detect_date_formats(rows, data_start_row_index)

    return {
        "file_type": file_type,
        "sheet_name": resolved_sheet,
        "header_row_index": header_row_index,
        "data_start_row_index": data_start_row_index,
        "field_aliases": field_aliases,
        "amount_mode": amount_mode.value if hasattr(amount_mode, "value") else amount_mode,
        "amount_config": amount_config,
        "date_formats": date_formats,
    }


def detect_journal_template_config(
    path: str | Path,
    file_type: str,
    sheet_name: str = "",
) -> dict[str, object]:
    """从日记账样本文件自动识别模板配置。

    比银行模板 detect 更简单：日记账列名就是输出列名本身（不映射到标准字段），
    无需字段别名/金额模式/日期格式。只需：
    1. 识别表头行（用日记账关键词 JOURNAL_HEADER_KEYWORDS 打分）
    2. 取表头单元格作为 columns（非空单元格即列名）
    3. 推断默认必填列（与识别列名取交集）
    返回可直接填入 CompanyJournalTemplateVersionCreate 的 dict。
    """
    resolved_sheet = sheet_name or _first_sheet_name(path, file_type)
    it = _read_rows(path, file_type, resolved_sheet)
    scan_limit = 30
    window = list(itertools.islice(it, scan_limit))
    header_row_index = detect_header_row(window, keywords=JOURNAL_HEADER_KEYWORDS)
    rows = list(itertools.chain(window, it))
    header_row = rows[header_row_index] if header_row_index < len(rows) else []
    data_start_row_index = _detect_data_start_row(rows, header_row_index)

    # 表头单元格原样取为列名（非空单元格）
    columns = [str(_clean_cell(cell)) for cell in header_row if _clean_cell(cell)]
    # 默认必填列：与识别出的列名取交集（样本没有的不强加）
    required_columns = [c for c in columns if c in JOURNAL_DEFAULT_REQUIRED]

    return {
        "file_type": file_type,
        "sheet_name": resolved_sheet,
        "header_row_index": header_row_index,
        "data_start_row_index": data_start_row_index,
        "columns": columns,
        "required_columns": required_columns,
    }


def _first_sheet_name(path: str | Path, file_type: str) -> str:
    if file_type.lower() != "xlsx":
        return "Sheet1"
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames[0] if workbook.sheetnames else "Sheet1"
    finally:
        workbook.close()


def list_xlsx_sheets(path: str | Path) -> list[str]:
    """返回 xlsx 文件的全部工作表名（按文件内顺序）。

    供「上传后让用户选择工作表」能力使用。非 xlsx / 文件不存在 / 解析失败
    均返回空列表（不抛错），调用方据此对 CSV 隐藏工作表选择器。

    read_only 模式只读元数据，不物化单元格，开销很低。
    """
    file_path = Path(path)
    if not file_path.exists():
        return []
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        # openpyxl 对非 xlsx（如 CSV）会抛 InvalidFileException 等；统一降级为空
        return []
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _detect_data_start_row(rows: list[list[CellValue]], header_row_index: int) -> int:
    """表头之后第一个非空数据行。"""
    for index in range(header_row_index + 1, len(rows)):
        if not _row_is_empty(rows[index]):
            return index
    return header_row_index + 1


def _detect_field_aliases(
    header_row: list[CellValue],
    custom_fields: list[CustomFieldDef] | None = None,
    builtin_keyword_overrides: dict[str, list[str]] | None = None,
) -> dict[str, str]:
    """把表头中文关键词映射到标准字段名（含公司级扩展字段 + 内置关键词覆盖）。"""
    aliases: dict[str, str] = {}
    used_fields: set[str] = set()
    for cell in header_row:
        header = _clean_cell(cell)
        if not header:
            continue
        if header in aliases:
            continue
        canonical = _match_canonical_field(header, custom_fields, builtin_keyword_overrides)
        if canonical is not None and canonical not in used_fields:
            aliases[header] = canonical
            used_fields.add(canonical)
    return aliases


def _match_canonical_field(
    header: str,
    custom_fields: list[CustomFieldDef] | None = None,
    builtin_keyword_overrides: dict[str, list[str]] | None = None,
) -> str | None:
    """返回表头对应的标准字段名（精确优先，再前缀匹配）。

    匹配顺序：
    1. 内置 HEADER_FIELD_MAP（可被 builtin_keyword_overrides 覆盖关键词集）
    2. 公司级扩展字段的 header_keywords

    builtin_keyword_overrides: {field_key: 完整关键词列表（含内置默认+覆盖）}。
    传入后，该 field_key 的识别用覆盖后的关键词集，否则用 HEADER_FIELD_MAP 默认。
    """
    # 内置字段：优先用覆盖后的关键词集（若有），否则用 HEADER_FIELD_MAP
    if builtin_keyword_overrides:
        for canonical, keywords in builtin_keyword_overrides.items():
            if header in keywords:
                return canonical
    for keywords, canonical in HEADER_FIELD_MAP.items():
        if builtin_keyword_overrides and canonical in builtin_keyword_overrides:
            continue  # 已用覆盖集匹配过，跳过默认集避免重复
        if header in keywords:
            return canonical
    # 前缀/子串匹配（同样优先覆盖集）
    if builtin_keyword_overrides:
        for canonical, keywords in builtin_keyword_overrides.items():
            if any(header.startswith(k) or k in header for k in keywords):
                return canonical
    for keywords, canonical in HEADER_FIELD_MAP.items():
        if builtin_keyword_overrides and canonical in builtin_keyword_overrides:
            continue
        if any(header.startswith(keyword) or keyword in header for keyword in keywords):
            return canonical
    # 公司级扩展字段：header_keywords → field_key
    if custom_fields:
        for cf in custom_fields:
            kws = cf.header_keywords or []
            if header in kws:
                return cf.field_key
        for cf in custom_fields:
            kws = cf.header_keywords or []
            if any(header.startswith(k) or k in header for k in kws):
                return cf.field_key
    return None


def _detect_amount_mode(
    field_aliases: dict[str, str],
) -> tuple[AmountMode, dict[str, str]]:
    """根据识别出的字段推断金额模式与 amount_config。"""
    values = set(field_aliases.values())
    if "income_amount" in values and "expense_amount" in values:
        return AmountMode.INCOME_EXPENSE_COLUMNS, {
            "income": "income_amount",
            "expense": "expense_amount",
        }
    if "debit_amount" in values and "credit_amount" in values:
        return AmountMode.DEBIT_CREDIT_COLUMNS, {
            "debit": "debit_amount",
            "credit": "credit_amount",
        }
    if "amount" in values and "direction" in values:
        return AmountMode.SINGLE_AMOUNT_WITH_DIRECTION, {
            "amount": "amount",
            "direction": "direction",
        }
    if "amount" in values:
        return AmountMode.SIGNED_AMOUNT, {"amount": "amount"}
    # 兜底：无法识别时默认 income_expense_columns（保守）。
    return AmountMode.INCOME_EXPENSE_COLUMNS, {
        "income": "income_amount",
        "expense": "expense_amount",
    }


def _detect_date_formats(
    rows: list[list[CellValue]], data_start_row_index: int
) -> list[str]:
    """从首行数据探测日期格式，返回候选列表。"""
    candidates = ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%d/%m/%Y", "%m/%d/%Y"]
    sample = ""
    if data_start_row_index < len(rows):
        for cell in rows[data_start_row_index]:
            cleaned = _clean_cell(cell)
            if cleaned and any(c.isdigit() for c in cleaned):
                sample = cleaned
                break
    matched: list[str] = []
    for date_format in candidates:
        try:
            datetime.strptime(sample, date_format)
            matched.append(date_format)
        except ValueError:
            continue
    # 始终保留默认格式作为兜底
    if "%Y-%m-%d" not in matched:
        matched.insert(0, "%Y-%m-%d")
    return matched



def _read_csv_text(file_path: Path) -> str:
    """按优先级尝试解码:UTF-8(含 BOM)→ GB18030(GBK/GB2312 超集)。"""
    raw = file_path.read_bytes()
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode CSV: tried utf-8 and gb18030")


def _iter_csv_rows(file_path: Path) -> Iterator[list[CellValue]]:
    """流式迭代 CSV 行（惰性生成器，不全量物化行列表）。

    编码检测需完整读取字节（utf-8-sig 优先，gb18030 兜底），之后用 io.StringIO
    包装成文件对象，让 csv.reader 逐行惰性消费，避免 splitlines() 二次物化。
    StringIO 不持有文件句柄，无资源泄漏风险。
    """
    text = _read_csv_text(file_path)
    for row in csv.reader(io.StringIO(text)):
        yield [_clean_cell(cell) for cell in row]


def _iter_xlsx_rows(file_path: Path, sheet_name: str) -> Iterator[list[CellValue]]:
    """流式迭代 XLSX 行（read_only + iter_rows，不全量物化行列表）。

    workbook 在 finally 块关闭：生成器正常耗尽或被 close() 时均触发 finally，
    不存在句柄泄漏。
    """
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            yield list(row)
    finally:
        workbook.close()


def _read_rows(path: str | Path, file_type: str, sheet_name: str) -> Iterator[list[CellValue]]:
    """返回文件行的惰性迭代器（不全量物化）。调用方按需消费或 list() 物化。"""
    file_path = Path(path)
    normalized_type = file_type.lower()

    if normalized_type == "csv":
        return _iter_csv_rows(file_path)

    if normalized_type == "xlsx":
        return _iter_xlsx_rows(file_path, sheet_name)

    if normalized_type == "xls":
        raise ValueError("Unsupported file type: xls")

    raise ValueError(f"Unsupported file type: {file_type}")


def _build_raw_row(header_row: list[CellValue], row: list[CellValue]) -> dict[str, CellValue]:
    raw_row: dict[str, CellValue] = {}

    for index, header in enumerate(header_row):
        key = _clean_cell(header)
        if not key:
            continue
        raw_row[key] = row[index] if index < len(row) else None

    return raw_row


def _normalize_row(
    raw_row: dict[str, CellValue],
    field_aliases: dict[str, str],
) -> dict[str, CellValue]:
    normalized: dict[str, CellValue] = {}

    for header, value in raw_row.items():
        canonical_field = field_aliases.get(header)
        if canonical_field is None:
            continue
        normalized[canonical_field] = value

    return normalized


DIRECTION_KEYWORDS_CREDIT = {"收入", "贷", "贷方", "入", "credit", "income"}
DIRECTION_KEYWORDS_DEBIT = {"支出", "借", "借方", "出", "debit", "expense"}

_FULLWIDTH_MAP = str.maketrans(
    "０１２３４５６７８９．，（）　",
    "0123456789.,() ",
)


def _parse_amounts(
    normalized_row: dict[str, CellValue],
    amount_mode: AmountMode,
    amount_config: dict[str, str],
) -> SignedAmount:
    if amount_mode == AmountMode.INCOME_EXPENSE_COLUMNS:
        return SignedAmount.from_income_expense(
            _decimal_or_none(normalized_row.get(amount_config["income"])),
            _decimal_or_none(normalized_row.get(amount_config["expense"])),
        )
    if amount_mode == AmountMode.DEBIT_CREDIT_COLUMNS:
        return SignedAmount.from_debit_credit(
            _decimal_or_none(normalized_row.get(amount_config["debit"])),
            _decimal_or_none(normalized_row.get(amount_config["credit"])),
        )
    if amount_mode == AmountMode.SINGLE_AMOUNT_WITH_DIRECTION:
        amount = _decimal_or_none(normalized_row.get(amount_config["amount"]))
        if amount is None:
            raise AmountError("Unable to determine transaction amount")
        direction_raw = _clean_cell(normalized_row.get(amount_config["direction"]))
        if not direction_raw:
            raise AmountError("Missing direction for single amount mode")
        direction = _direction_from_keyword(direction_raw)
        if direction is None:
            raise AmountError(f"Unknown direction value: {direction_raw}")
        return SignedAmount.from_amount_with_direction(amount, direction)
    if amount_mode == AmountMode.SIGNED_AMOUNT:
        amount = _decimal_or_none(normalized_row.get(amount_config["amount"]))
        if amount is None:
            raise AmountError("Unable to determine transaction amount")
        return SignedAmount.from_signed(amount)
    raise ValueError(f"Unsupported amount mode: {amount_mode}")


def _direction_from_keyword(raw: str) -> TransactionDirection | None:
    if raw in DIRECTION_KEYWORDS_CREDIT:
        return TransactionDirection.CREDIT
    if raw in DIRECTION_KEYWORDS_DEBIT:
        return TransactionDirection.DEBIT
    # 容忍 "存入/支出"、"贷方/借方" 等包含关键词的写法
    for keyword in DIRECTION_KEYWORDS_CREDIT:
        if keyword in raw:
            return TransactionDirection.CREDIT
    for keyword in DIRECTION_KEYWORDS_DEBIT:
        if keyword in raw:
            return TransactionDirection.DEBIT
    return None


def _parse_date(value: CellValue, date_formats: list[str]) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    candidate = _clean_cell(value)
    for date_format in date_formats:
        try:
            return datetime.strptime(candidate, date_format).date().isoformat()
        except ValueError:
            continue

    raise ValueError(f"Invalid date: {value}")


def _parse_optional_date(value: CellValue, date_formats: list[str]) -> str | None:
    if value is None or not _clean_cell(value):
        return None
    return _parse_date(value, date_formats)


def _decimal_or_none(value: CellValue) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    candidate = _clean_cell(value)
    if not candidate:
        return None

    # 全角 → 半角(数字、逗号、句点、括号、空格)
    candidate = candidate.translate(_FULLWIDTH_MAP)
    negative = False
    # 会计括号负数:(1,000) / （1,000）
    if candidate.startswith("(") and candidate.endswith(")"):
        negative = True
        candidate = candidate[1:-1]
    # 方向后缀 DR/CR
    upper = candidate.upper()
    if upper.endswith("DR"):
        negative = True
        candidate = candidate[:-2]
    elif upper.endswith("CR"):
        candidate = candidate[:-2]
    # 去货币符号、千分位、空白
    for token in ("¥", "￥", "$", "CNY", "RMB", ",", " "):
        candidate = candidate.replace(token, "")
    if not candidate:
        return None

    try:
        result = Decimal(candidate)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid amount: {value}") from exc
    return -result if negative else result


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _none_if_blank(value: CellValue) -> str | None:
    if value is None:
        return None
    cleaned = _clean_cell(value)
    return cleaned or None


def _row_is_empty(row: list[CellValue]) -> bool:
    return all(not _clean_cell(cell) for cell in row)
