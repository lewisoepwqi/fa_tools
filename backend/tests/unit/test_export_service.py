from pathlib import Path

from openpyxl import load_workbook

from app.tools.bank_journal.services.export_service import (
    export_preview_rows_to_csv,
    export_preview_rows_to_xlsx,
)


def test_export_preview_rows_to_csv(tmp_path: Path) -> None:
    rows = [
        {"日期": "2026-06-01", "摘要": "收到客户款项", "科目": "银行存款", "金额": "12000.00"},
        {"日期": "2026-06-02", "摘要": "支付采购款", "科目": "管理费用", "金额": "-3000.00"},
    ]

    output = export_preview_rows_to_csv(
        rows=rows,
        columns=["日期", "摘要", "科目", "金额"],
        output_dir=tmp_path,
        filename="journal.csv",
    )

    assert output.read_text(encoding="utf-8-sig").splitlines()[0] == "日期,摘要,科目,金额"


def test_export_preview_rows_to_xlsx_writes_columns_and_rows(tmp_path: Path) -> None:
    rows = [{"日期": "2026-06-01", "摘要": "收到客户款项", "科目": "银行存款", "金额": "12000.00"}]
    output = export_preview_rows_to_xlsx(
        rows=rows,
        columns=["日期", "摘要", "科目", "金额"],
        output_dir=tmp_path,
        filename="journal.xlsx",
    )
    workbook = load_workbook(output)
    sheet = workbook.active
    assert [c.value for c in sheet[1]] == ["日期", "摘要", "科目", "金额"]
    assert sheet.cell(2, 1).value == "2026-06-01"
    assert sheet.cell(2, 3).value == "银行存款"
